import pandas as pd
from pathlib import Path
from typing import Dict
from io import StringIO
from app.utils.encoding import detect_encoding
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import logging

logger = logging.getLogger(__name__)


class ExcelConverter:
    """Handle Excel and CSV file conversions"""
    
    @staticmethod
    def apply_excel_styling(
        ws,
        header_row: int = 1,
        freeze_panes_cell: str = "A2",
        table_name: str = "DataTable",
        table_style: str = "TableStyleMedium9",
        min_row_height: float = 18,
        header_row_height: float = 26,
        wrap_text: bool = True,
    ):
        """
        Apply professional Excel styling:
        - Freeze header row
        - Auto-filter with table style
        - Auto-adjust column widths
        - Text wrapping + top alignment
        - Header emphasis
        - Borders
        - Minimum row height
        """
        # 1) Freeze panes
        ws.freeze_panes = freeze_panes_cell

        # 2) Determine used range
        max_row = ws.max_row
        max_col = ws.max_column
        
        if max_row < header_row or max_col < 1:
            return  # No data to style

        # 3) Styles
        header_fill = PatternFill("solid", fgColor="1F4E79")  # Dark blue
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

        body_alignment = Alignment(
            vertical="top",
            horizontal="left",
            wrap_text=wrap_text
        )

        thin = Side(style="thin", color="9E9E9E")
        border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

        # 4) Apply header style
        ws.row_dimensions[header_row].height = header_row_height
        for c in range(1, max_col + 1):
            cell = ws.cell(row=header_row, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border_all

        # 5) Apply body style + borders + minimum row height
        for r in range(header_row + 1, max_row + 1):
            current_height = ws.row_dimensions[r].height or 0
            ws.row_dimensions[r].height = max(current_height, min_row_height)
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.alignment = body_alignment
                cell.border = border_all

        # 6) Auto-fit column widths
        max_width_cap = 60
        min_width_floor = 10

        for c in range(1, max_col + 1):
            max_len = 0
            for r in range(1, max_row + 1):
                v = ws.cell(row=r, column=c).value
                if v is None:
                    continue
                s = str(v)
                # Handle multi-line text
                longest_line = max((len(line) for line in s.splitlines()), default=len(s))
                max_len = max(max_len, longest_line)
            
            # Add padding
            width = min(max_width_cap, max(min_width_floor, int(max_len * 1.1) + 2))
            ws.column_dimensions[get_column_letter(c)].width = width

        # 7) Create Excel Table for AutoFilter + banded rows
        start_cell = f"A{header_row}"
        end_cell = f"{get_column_letter(max_col)}{max_row}"
        table_ref = f"{start_cell}:{end_cell}"

        # Avoid duplicate table names
        existing_names = {t.name for t in ws._tables}
        original_table_name = table_name
        i = 2
        while table_name in existing_names:
            table_name = f"{original_table_name}{i}"
            i += 1

        tab = Table(displayName=table_name, ref=table_ref)
        tab.tableStyleInfo = TableStyleInfo(
            name=table_style,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(tab)

        # 8) Enable auto-filter
        ws.auto_filter.ref = table_ref
    
    @staticmethod
    def excel_to_csv_dict(file_path: Path) -> Dict[str, str]:
        """
        Convert Excel file to dictionary of CSV strings
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            Dict with sheet names as keys and CSV strings as values
        """
        result = {}
        
        # Read Excel file
        excel_file = pd.ExcelFile(file_path)
        
        for sheet_name in excel_file.sheet_names:
            df = pd.read_excel(excel_file, sheet_name=sheet_name)
            
            # Convert to CSV string
            csv_buffer = StringIO()
            df.to_csv(csv_buffer, index=False, encoding='utf-8')
            result[sheet_name] = csv_buffer.getvalue()
        
        return result
    
    @staticmethod
    def csv_to_excel(csv_dict: Dict[str, str], output_path: Path) -> None:
        """
        Convert dictionary of CSV strings to Excel file with professional styling
        
        Args:
            csv_dict: Dict with sheet names as keys and CSV strings as values
            output_path: Path to save Excel file
        """
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for idx, (sheet_name, csv_content) in enumerate(csv_dict.items(), 1):
                try:
                    # Read CSV with proper handling of quotes and delimiters
                    df = pd.read_csv(
                        StringIO(csv_content),
                        quotechar='"',
                        escapechar='\\',
                        on_bad_lines='skip',  # Skip malformed lines instead of warning
                        engine='python',
                        encoding='utf-8'
                    )
                    
                    # Log if rows were skipped
                    expected_rows = csv_content.count('\n')
                    actual_rows = len(df)
                    if actual_rows < expected_rows - 1:
                        logger.warning(f"Sheet '{sheet_name}': {expected_rows - actual_rows - 1} rows skipped due to CSV format issues")
                    
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                except Exception as e:
                    logger.error(f"Error converting sheet '{sheet_name}': {str(e)}")
                    # Create empty dataframe as fallback
                    df = pd.DataFrame()
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Apply styling after all sheets are written
        wb = load_workbook(output_path)
        for idx, sheet_name in enumerate(csv_dict.keys(), 1):
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                ExcelConverter.apply_excel_styling(
                    ws,
                    header_row=1,
                    freeze_panes_cell="A2",
                    table_name=f"Table{idx}",
                    table_style="TableStyleMedium9"
                )
        wb.save(output_path)
    
    @staticmethod
    def csv_file_to_string(file_path: Path) -> str:
        """
        Read CSV file and return as string
        
        Args:
            file_path: Path to CSV file
            
        Returns:
            CSV content as string
        """
        encoding, _ = detect_encoding(file_path)
        
        with open(file_path, 'r', encoding=encoding) as f:
            return f.read()
    
    @staticmethod
    def csv_string_to_file(csv_content: str, output_path: Path) -> None:
        """
        Write CSV string to file
        
        Args:
            csv_content: CSV content as string
            output_path: Path to save CSV file
        """
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(csv_content)

# Made with Bob
